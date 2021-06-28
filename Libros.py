import openpyxl

class Books:
    All_Books = []

    @staticmethod
    def full_list(path):
        Books.All_Books = []
        # Leer el documento con openpyxl
        document = openpyxl.load_workbook(path)

        # Crear el objeto sheet del excel
        sheet_obj = document.active

        # Columnas y filas maximas del documento. Esto para que se pueda agregar estudiantes
        max_row = sheet_obj.max_row + 1
        max_column = sheet_obj.max_column + 1

        # Recorrer el archivo por filas, comenzando en la fila 2
        row_start = 2
        for row_i in range(row_start, max_row):
            column_start = 1
            # Lista temporal a ser agregada
            temp_list = []
            # Obtener el valor de las columnas y meterlos en lista
            for column_i in range(column_start, max_column):
                cell_obj = sheet_obj.cell(row=row_i, column=column_i)
                if (column_i == 1):
                    temp_list.append((cell_obj.value))
                else:
                    temp_list.append(str(cell_obj.value))

            #Cantidad de libros en existencia
            amount = 10
            #Agregar a la variable global el libro
            Books.All_Books.append(temp_list + [amount])

        document.close()

    @staticmethod
    def to_String():
        string = ""
        for book in Books.All_Books:
            string += "Id: " + str(book[0]) + " || Nombre: " + book[1]+ " || Genero: " + book[2]+ " || Autor: " + book[3] + "\n"
        return string

    @staticmethod
    def get_by_name_gener_autor(name, gender, autor):
        for book in Books.All_Books:
            if((name == book[1]) and (gender == book[2]) and (autor == book[3])):
                return book
        return None

    @staticmethod
    def get_by_code(code):
        for book in Books.All_Books:
            if(code == book[0]):
                return book
        return None
