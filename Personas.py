import openpyxl

class Person:
    All_Person = []

    @staticmethod
    def full_list(path):
        Person.All_Person = []
        # Leer el documento con openpyxl
        document = openpyxl.load_workbook(path)

        # Crear el objeto sheet del excel
        sheet_obj = document.active

        # Columnas y filas maximas del documento. Esto para que se pueda agregar estudiantes
        max_row = sheet_obj.max_row + 1
        max_column = sheet_obj.max_column + 1

        # Recorrer el archivo por filas, comenzando en la fila 2
        row_start = 2
        for row_i in range(row_start, max_row):
            column_start = 1
            #Lista temporal a ser agregada
            temp_list = []
            # Obtener el valor de las columnas y meterlos en lista
            for column_i in range(column_start, max_column):
                cell_obj = sheet_obj.cell(row=row_i, column=column_i)
                if(column_i == 1):
                    temp_list.append((cell_obj.value))
                else:
                    temp_list.append(str(cell_obj.value))

            #Lista con codigo de libros prestados
            books = []
            #Agregar persona a la lista global
            Person.All_Person.append(temp_list + [books])

        document.close()

    @staticmethod
    def to_String():
        string = ""
        for person in Person.All_Person:
            string += "Codigo: " + str(person[0]) + " || Nombre: " + person[1] + " || Correo: " + person[2] + "\n"
        return string

    @staticmethod
    def get_by_name(name):
        for person in Person.All_Person:
            if(name == person[1]):
                return person
        return None

    @staticmethod
    def get_by_code(code):
        for person in Person.All_Person:
            if(code == person[0]):
                return person
        return None